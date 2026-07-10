"""
Core logic for the Billables workflow:
  1. Parse a weekly "Billables" Word document into structured entries.
  2. Write those entries into a new tab of the master workbook.
  3. Consolidate all "Week N" tabs into a Consolidated tab + Summary tab.

Kept separate from app.py (the Streamlit UI) so it can be tested/run
without needing Streamlit installed.
"""

import re
from datetime import datetime
from io import BytesIO

from docx import Document
from docx.oxml.ns import qn
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

DATE_RE = re.compile(r'^(\d{2}/\d{2}/\d{2})\s+(.*)$', re.S)
AMOUNT_RE = re.compile(r'\$[\d,]+\.\d{2}')
WEEK_SHEET_RE = re.compile(r'^Week\s*(\d+)$', re.IGNORECASE)


# ---------------------------------------------------------------------------
# STEP 1: Parse the weekly Word document
# ---------------------------------------------------------------------------

def _iter_block_items(parent):
    """Yield paragraphs and tables in the order they appear in the document body."""
    body = parent.element.body
    for child in body.iterchildren():
        if child.tag == qn('w:p'):
            from docx.text.paragraph import Paragraph
            yield Paragraph(child, parent)
        elif child.tag == qn('w:tbl'):
            from docx.table import Table
            yield Table(child, parent)


def _paragraph_highlighted_and_other_text(paragraph):
    """
    Split a paragraph's text into (highlighted_text, other_text) based on each
    run's highlight formatting. Word's manual line breaks show up as literal
    '\\n' inside run.text; those are collapsed to spaces within each bucket so
    a visually-wrapped sentence stays one continuous string.
    """
    highlighted_parts = []
    other_parts = []
    for run in paragraph.runs:
        text = run.text
        if not text:
            continue
        if run.font.highlight_color is not None:
            highlighted_parts.append(text)
        else:
            other_parts.append(text)

    def clean(parts):
        joined = ''.join(parts).replace('\n', ' ').replace('\r', ' ')
        return re.sub(r'\s+', ' ', joined).strip()

    return clean(highlighted_parts), clean(other_parts)


def _entry_from_line(text):
    """If a line is a billable entry (starts with a date), parse it. Else None."""
    m = DATE_RE.match(text.strip())
    if not m:
        return None
    date_str, rest = m.groups()
    rest = rest.strip()
    matches = list(AMOUNT_RE.finditer(rest))
    if not matches:
        return None
    last = matches[-1]
    amount_str = last.group()
    description = (rest[:last.start()] + rest[last.end():]).strip()
    # tidy up stray punctuation/space left behind where the amount was removed
    description = re.sub(r'\s{2,}', ' ', description).strip()
    try:
        date_val = datetime.strptime(date_str, '%m/%d/%y')
    except ValueError:
        date_val = date_str
    amount_val = float(amount_str.replace('$', '').replace(',', ''))
    return date_val, description, amount_val


def parse_billables_docx(file_obj):
    """
    Parse a weekly Billables .docx file.
    Returns a list of (property_name, date, description, amount) tuples,
    in the same order they appear in the document.
    """
    doc = Document(file_obj)
    entries = []

    property_buffer = []
    in_property_block = False
    current_property = ''

    def all_paragraphs():
        for block in _iter_block_items(doc):
            if block.__class__.__name__ == 'Paragraph':
                yield block
            else:  # Table
                for row in block.rows:
                    for cell in row.cells:
                        for p in cell.paragraphs:
                            yield p

    for paragraph in all_paragraphs():
        hl_text, other_text = _paragraph_highlighted_and_other_text(paragraph)

        if hl_text:
            if not in_property_block:
                property_buffer = []
                in_property_block = True
            property_buffer.append(hl_text)

        if other_text:
            if in_property_block:
                current_property = ' '.join(property_buffer).strip()
                in_property_block = False

            parsed = _entry_from_line(other_text)
            if parsed:
                date_val, description, amount_val = parsed
                entries.append((current_property, date_val, description, amount_val))
            # else: technician name / title line / other non-entry text -> ignore

    return entries


# ---------------------------------------------------------------------------
# STEP 2: Write entries into a new "Week N" tab
# ---------------------------------------------------------------------------

HEADER_FONT = Font(bold=True, name='Arial', size=11)
BODY_FONT = Font(name='Arial', size=10)
TOTAL_FONT = Font(bold=True, name='Arial', size=10)
WRAP = Alignment(wrap_text=True, vertical='top')

COL_WIDTHS = {'A': 32, 'B': 12, 'C': 90, 'D': 14}


def _write_entries_sheet(ws, entries, include_week_col=False):
    headers = (['Week'] if include_week_col else []) + ['Property Name', 'Date', 'Description', 'Amount']
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = HEADER_FONT

    offset = 1 if include_week_col else 0
    r = 2
    amount_rows = []
    for entry in entries:
        if include_week_col:
            week_label, prop, date_val, desc, amt = entry
            ws.cell(row=r, column=1, value=week_label).font = BODY_FONT
        else:
            prop, date_val, desc, amt = entry

        ws.cell(row=r, column=1 + offset, value=prop).font = BODY_FONT
        date_cell = ws.cell(row=r, column=2 + offset, value=date_val)
        date_cell.font = BODY_FONT
        date_cell.number_format = 'MM/DD/YY'
        desc_cell = ws.cell(row=r, column=3 + offset, value=desc)
        desc_cell.font = BODY_FONT
        desc_cell.alignment = WRAP
        amt_cell = ws.cell(row=r, column=4 + offset, value=amt)
        amt_cell.font = BODY_FONT
        amt_cell.number_format = '$#,##0.00'
        amount_rows.append(r)
        r += 2  # blank row between entries

    total_row = r + 1
    ws.cell(row=total_row, column=3 + offset, value='Total').font = TOTAL_FONT
    if amount_rows:
        formula = '=' + '+'.join(f'{get_column_letter(4 + offset)}{row}' for row in amount_rows)
    else:
        formula = 0
    total_cell = ws.cell(row=total_row, column=4 + offset, value=formula)
    total_cell.font = TOTAL_FONT
    total_cell.number_format = '$#,##0.00'

    letters = (['A', 'B', 'C', 'D', 'E'] if include_week_col else ['A', 'B', 'C', 'D'])
    base_widths = ([10] if include_week_col else []) + [32, 12, 90, 14]
    for letter, width in zip(letters, base_widths):
        ws.column_dimensions[letter].width = width

    return total_row


def add_week_sheet(wb, sheet_name, entries):
    """Add (or replace) a sheet named sheet_name in wb containing entries."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    _write_entries_sheet(ws, entries, include_week_col=False)
    return ws


def load_or_create_workbook(file_obj=None):
    if file_obj is not None:
        wb = load_workbook(file_obj)
        # drop the default empty "Sheet" if present and unused
        if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
            ws = wb['Sheet']
            if ws.max_row == 1 and ws.max_column == 1 and ws['A1'].value is None:
                del wb['Sheet']
        return wb
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def next_week_sheet_name(wb):
    """Suggest the next 'Week N' name based on existing sheets."""
    nums = [int(m.group(1)) for name in wb.sheetnames if (m := WEEK_SHEET_RE.match(name))]
    n = (max(nums) + 1) if nums else 1
    return f'Week {n}'


# ---------------------------------------------------------------------------
# STEP 3: Consolidate all Week N sheets into Consolidated + Summary tabs
# ---------------------------------------------------------------------------

def _read_entries_from_week_sheet(ws):
    """Re-read a Week N sheet's data rows (skips header, blanks, total row)."""
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        prop_cell, date_cell, desc_cell, amt_cell = row[0], row[1], row[2], row[3]
        if desc_cell.value == 'Total':
            break
        if prop_cell.value is None and date_cell.value is None:
            continue
        entries.append((prop_cell.value, date_cell.value, desc_cell.value, amt_cell.value))
    return entries


def consolidate_workbook(wb):
    """
    Build/replace a 'Consolidated' sheet (all entries from every Week N sheet,
    tagged by week) and a 'Summary' sheet (SUMIF total per property).
    Returns (consolidated_row_count, summary_row_count).
    """
    week_sheets = sorted(
        (name for name in wb.sheetnames if WEEK_SHEET_RE.match(name)),
        key=lambda n: int(WEEK_SHEET_RE.match(n).group(1))
    )
    if not week_sheets:
        raise ValueError('No sheets named "Week 1", "Week 2", etc. were found in this workbook.')

    all_entries = []
    for name in week_sheets:
        for prop, date_val, desc, amt in _read_entries_from_week_sheet(wb[name]):
            all_entries.append((name, prop, date_val, desc, amt))

    def sort_key(entry):
        _, prop, date_val, _, _ = entry
        prop_key = (prop or '').strip().lower()
        # dates are normally real datetime objects; fall back gracefully if not
        date_key = date_val if hasattr(date_val, 'toordinal') else datetime.min
        return (prop_key, date_key)

    all_entries.sort(key=sort_key)

    if 'Consolidated' in wb.sheetnames:
        del wb['Consolidated']
    cons_ws = wb.create_sheet('Consolidated', 0)
    _write_entries_sheet(cons_ws, all_entries, include_week_col=True)

    # ---- Summary by property (SUMIF against the Consolidated tab) ----
    properties = []
    seen = set()
    for _, prop, _, _, _ in all_entries:
        if prop not in seen:
            seen.add(prop)
            properties.append(prop)

    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    sum_ws = wb.create_sheet('Summary', 1)
    sum_ws.append(['Property Name', 'Total Amount'])
    sum_ws['A1'].font = HEADER_FONT
    sum_ws['B1'].font = HEADER_FONT

    r = 2
    for prop in properties:
        sum_ws.cell(row=r, column=1, value=prop).font = BODY_FONT
        formula = f'=SUMIF(Consolidated!B:B,A{r},Consolidated!E:E)'
        amt_cell = sum_ws.cell(row=r, column=2, value=formula)
        amt_cell.font = BODY_FONT
        amt_cell.number_format = '$#,##0.00'
        r += 1

    grand_row = r + 1
    sum_ws.cell(row=grand_row, column=1, value='Grand Total').font = TOTAL_FONT
    grand_cell = sum_ws.cell(row=grand_row, column=2, value=f'=SUM(B2:B{r - 1})')
    grand_cell.font = TOTAL_FONT
    grand_cell.number_format = '$#,##0.00'

    sum_ws.column_dimensions['A'].width = 40
    sum_ws.column_dimensions['B'].width = 16

    return len(all_entries), len(properties)


def workbook_to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
